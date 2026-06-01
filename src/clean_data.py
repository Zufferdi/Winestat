"""
clean_data.py — Read the source xlsx and emit tidy CSV files into data/.

Source: Office fédéral de l'agriculture (OFAG / BLW) — Contingents
d'importation de vin 2025, plus historique OFDF 2012-2026.
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

SRC = Path("data/raw/Contingents_d_importation_de_vin_2025.xlsx")
OUT = Path("data/processed")
OUT.mkdir(parents=True, exist_ok=True)


def clean_importers() -> pd.DataFrame:
    """Master table of importers from the 'Full' sheet."""
    df = pd.read_excel(SRC, sheet_name="Full")

    df = df.rename(columns={
        "IMPORTATEUR": "importateur",
        "NPA": "npa",
        "COMMUNE": "commune",
        "Canton": "canton",
        "(rouge) Litres": "litres_rouge",
        "(blanc) litres": "litres_blanc",
        "Total": "litres_total",
        "NOGA": "noga_code",
        "Intitulé": "noga_label",
        "Commentaire": "commentaire",
        "Groupe": "groupe",
    })

    df["litres_rouge"] = df["litres_rouge"].fillna(0).astype(float)
    df["litres_blanc"] = df["litres_blanc"].fillna(0).astype(float)
    df["litres_total"] = df["litres_rouge"] + df["litres_blanc"]
    df["npa"] = df["npa"].astype("Int64")

    df["pays"] = df["canton"].apply(lambda c: "LI" if c == "FL" else "CH")
    df = df.sort_values("litres_total", ascending=False).reset_index(drop=True)
    df.to_csv(OUT / "importateurs.csv", index=False)
    return df


def clean_history() -> pd.DataFrame:
    """OFDF historical series (kilos, CHF, price/kg) by HS code section."""
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Importation (OFDF)"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_title = None
    for r in rows:
        if r[0] is None and r[1] is None:
            continue
        if isinstance(r[0], str):
            current_title = r[0]
        elif isinstance(r[0], (int, float)) and r[1] is not None:
            records.append({
                "section": current_title,
                "annee": int(r[0]),
                "kilos": float(r[1]),
                "chf": float(r[2]),
                "prix_par_kg": float(r[3]) if r[3] is not None else None,
            })

    df = pd.DataFrame(records)

    def short(s: str) -> str:
        if s.startswith("2204.2121"):
            return "Vin blanc (≤2L)"
        if s.startswith("2204.2141"):
            return "Vin rouge (≤2L)"
        return "Tous vins"
    df["section_courte"] = df["section"].apply(short)
    df["complet"] = df["annee"] < 2026

    df.to_csv(OUT / "historique_ofdf.csv", index=False)
    return df


def main():
    importers = clean_importers()
    history = clean_history()
    print(f"Importateurs : {len(importers)} lignes")
    print(f"  Litres rouge totaux : {importers['litres_rouge'].sum():>15,.0f}")
    print(f"  Litres blanc totaux : {importers['litres_blanc'].sum():>15,.0f}")
    print(f"  Cantons couverts    : {importers['canton'].nunique()}")
    print(f"  Communes couvertes  : {importers['commune'].nunique()}")
    print(f"Historique OFDF      : {len(history)} lignes, "
          f"{history['annee'].min()}-{history['annee'].max()}")


if __name__ == "__main__":
    main()
